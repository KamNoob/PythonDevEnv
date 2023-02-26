import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

excel_file = pd.read_excel('supermarket_sales.xlsx')
excel_file[['Gender', 'Product line', 'Total']]

report_table = excel_file.pivot_table(index='Gender',columns='Product line', values='Total', aggfunc='sum').round(0)

report_table.to_excel('report_2021.xlsx', sheet_name='Report', startrow=3)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# cell redferences (original spreadsheet)
minColumn = wb.active.min_column
maxColumn = wb.active.max_column
minRow = wb.active.min_row
maxRow = wb.active.max_row

sheet['B7'].value = '=SUM(B5:B6)'
sheet['B7'].style =  'Currency'

wb.save('report_2021.xlsx')
wb.close()